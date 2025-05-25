import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook
from openpyxl.styles import Font
import traceback
import datetime

class StorageCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Расчет площади склада заполнителей (по формуле S = V/(q·Kис))")
        self.root.geometry("800x600")
        
        # Переменные для хранения данных
        self.materials = ["Щебень", "Керамзит", "Песок"]
        self.storage_types = {
            "Открытый склад штабельного типа": {"q_range": (3, 4), "height_range": (5, 6)},
            "Другие типы складов": {"q_range": (5, 7), "height_range": None}
        }
        
        self.create_widgets()
    
    def create_widgets(self):
        # Основной контейнер
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Вкладки для разных материалов
        self.notebook = ttk.Notebook(main_frame)
        
        self.material_frames = {}
        self.volume_vars = {}
        self.storage_type_vars = {}
        self.q_vars = {}
        self.kis_vars = {}
        
        for material in self.materials:
            frame = ttk.Frame(self.notebook, padding="10")
            self.material_frames[material] = frame
            self.notebook.add(frame, text=material)
            
            # Заголовок
            ttk.Label(frame, text=f"Параметры для {material}", font=('Arial', 12, 'bold')).grid(row=0, column=0, columnspan=2, pady=5)
            
            # Объем материала
            ttk.Label(frame, text="Объем склада заполнителей V (м³):").grid(row=1, column=0, sticky=tk.W, pady=5)
            self.volume_vars[material] = tk.DoubleVar()
            ttk.Entry(frame, textvariable=self.volume_vars[material]).grid(row=1, column=1, pady=5)
            
            # Тип склада
            ttk.Label(frame, text="Тип склада:").grid(row=2, column=0, sticky=tk.W, pady=5)
            self.storage_type_vars[material] = tk.StringVar()
            storage_type_cb = ttk.Combobox(frame, textvariable=self.storage_type_vars[material], 
                                         values=list(self.storage_types.keys()))
            storage_type_cb.grid(row=2, column=1, pady=5)
            storage_type_cb.bind("<<ComboboxSelected>>", lambda e, mat=material: self.update_q_slider(mat))
            
            # Количество материала на 1 м² (q)
            ttk.Label(frame, text="Количество материала на 1 м² (q, м³/м²):").grid(row=3, column=0, sticky=tk.W, pady=5)
            self.q_vars[material] = tk.DoubleVar()
            self.q_sliders = {}
            self.q_sliders[material] = ttk.Scale(frame, from_=3, to=7, variable=self.q_vars[material], 
                                               command=lambda v, mat=material: self.update_q_value(mat))
            self.q_sliders[material].grid(row=3, column=1, pady=5, sticky=tk.EW)
            self.q_value_labels = {}
            self.q_value_labels[material] = ttk.Label(frame, text="3.0")
            self.q_value_labels[material].grid(row=3, column=2, padx=5)
            
            # Коэффициент использования площади (Кис)
            ttk.Label(frame, text="Коэффициент использования площади (Кис):").grid(row=4, column=0, sticky=tk.W, pady=5)
            self.kis_vars[material] = tk.DoubleVar(value=0.7)
            ttk.Scale(frame, from_=0.7, to=0.8, variable=self.kis_vars[material], 
                     command=lambda v, mat=material: self.update_kis_value(mat)).grid(row=4, column=1, pady=5, sticky=tk.EW)
            self.kis_value_labels = {}
            self.kis_value_labels[material] = ttk.Label(frame, text="0.70")
            self.kis_value_labels[material].grid(row=4, column=2, padx=5)
        
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Кнопки расчета
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill=tk.X, pady=10)
        
        ttk.Button(button_frame, text="Рассчитать", command=self.calculate).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Сохранить в Excel", command=self.save_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Очистить", command=self.clear_fields).pack(side=tk.LEFT, padx=5)
        
        # Результаты
        results_frame = ttk.LabelFrame(main_frame, text="Результаты расчета по формуле S = V/(q·Kис)", padding="10")
        results_frame.pack(fill=tk.BOTH, expand=True)
        
        self.results_text = tk.Text(results_frame, height=10, wrap=tk.WORD)
        self.results_text.pack(fill=tk.BOTH, expand=True)
        
        # Статус бар
        self.status_var = tk.StringVar(value="Готов к работе")
        ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN).pack(fill=tk.X, pady=5)
    
    def update_q_slider(self, material):
        storage_type = self.storage_type_vars[material].get()
        if storage_type in self.storage_types:
            q_min, q_max = self.storage_types[storage_type]["q_range"]
            self.q_sliders[material].config(from_=q_min, to=q_max)
            self.q_vars[material].set(q_min)
            self.update_q_value(material)
    
    def update_q_value(self, material):
        value = self.q_vars[material].get()
        self.q_value_labels[material].config(text=f"{value:.1f}")
    
    def update_kis_value(self, material):
        value = self.kis_vars[material].get()
        self.kis_value_labels[material].config(text=f"{value:.2f}")
    
    def calculate(self):
        try:
            results = []
            total_area = 0
            
            for material in self.materials:
                _v = self.volume_vars[material].get()
                if _v <= 0:
                    raise ValueError(f"Объем для {material} должен быть положительным числом")
                
                q = self.q_vars[material].get()
                kis = self.kis_vars[material].get()
                
                # Расчет по формуле S = V/(q·Kис)
                _s = _v / (q * kis)
                
                total_area += _s
                
                results.append({
                    "material": material,
                    "volume": _v,
                    "q": q,
                    "kis": kis,
                    "area": _s
                })
            
            # Формируем текст результатов
            result_text = "РЕЗУЛЬТАТЫ РАСЧЕТА (по формуле S = V/(q·Kис)):\n\n"
            for res in results:
                result_text += (
                    f"{res['material']}:\n"
                    f" - Объем склада (V): {res['volume']:.2f} м³\n"
                    f" - Количество на 1 м² (q): {res['q']:.1f} м³/м²\n"
                    f" - Коэффициент использования (Кис): {res['kis']:.2f}\n"
                    f" - Рассчитанная площадь: {res['area']:.2f} м²\n\n"
                )
            
            result_text += f"ОБЩАЯ ПЛОЩАДЬ СКЛАДА: {total_area:.2f} м²"
            
            self.results_text.delete(1.0, tk.END)
            self.results_text.insert(tk.END, result_text)
            self.status_var.set("Расчет выполнен успешно")
            
            # Сохраняем результаты для экспорта
            self.last_results = {
                "materials": results,
                "total_area": total_area
            }
            
        except Exception as e:
            messagebox.showerror("Ошибка", f'{e}\n{traceback.format_exc()}')
            self.status_var.set("Ошибка при расчете")
    
    def save_to_excel(self):
        if not hasattr(self, 'last_results'):
            messagebox.showwarning("Предупреждение", "Сначала выполните расчет")
            return
        
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"Расчет_площади_склада_{datetime.datetime.now().strftime(('%Y-%m-%d_%H-%M-%S'))}.xlsx",
                title="Выберите место для сохранения файла"
            )
            
            if not file_path:  # Отмена сохранения
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты расчета"
            headers = ["Материал", "Объем склада (м³)", "q (м³/м²)", "Кис", "Площадь (м²)"]
            
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            
            for row, res in enumerate(self.last_results['materials'], 2):
                ws.cell(row=row, column=1, value=res['material'])
                ws.cell(row=row, column=2, value=res['volume'])
                ws.cell(row=row, column=3, value=res['q'])
                ws.cell(row=row, column=4, value=res['kis'])
                ws.cell(row=row, column=5, value=res['area'])
            
            last_row = len(self.last_results['materials']) + 3
            ws.cell(row=last_row, column=1, value="ИТОГО:").font = Font(bold=True)
            ws.cell(row=last_row, column=5, value=self.last_results['total_area']).font = Font(bold=True)
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
            
            wb.save(file_path)
            
            messagebox.showinfo("Успех", f"Файл успешно сохранен:\n{file_path}")
            self.status_var.set(f"Результаты сохранены в {file_path}")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл:\n{str(e)}")
            self.status_var.set("Ошибка при сохранении файла")
    
    def clear_fields(self):
        for material in self.materials:
            self.volume_vars[material].set(0)
            self.storage_type_vars[material].set("")
            self.q_vars[material].set(3)
            self.kis_vars[material].set(0.7)
            self.update_q_value(material)
            self.update_kis_value(material)
        
        self.results_text.delete(1.0, tk.END)
        self.status_var.set("Поля очищены. Готов к работе")

if __name__ == "__main__":
    root = tk.Tk()
    app = StorageCalculatorApp(root)
    root.mainloop()
